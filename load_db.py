import os.path

from openpyxl import load_workbook

import create_db_from_file
from database.create_db_from_file import RRP, path
from engine import get_product_name


class FileProcess:
    def __init__(self, file_name):
        self.file_name = file_name
        self.ws = None
        self.dict = {}
        self.sig_cat = None
        self.categories = (
            "Манежи",
            "Стульчики",
            "автокресла",
            "Колыбели",
            "Коляски",
            "Базы",
            "SIGNATURE SERIES Кресла",
            "SIGNATURE SERIES Коляски",
        )
        self.categorie = None

    def main(self):
        self.wb_active()
        for i in range(3, self.ws.max_row + 1):
            name = self.ws[i][0].value
            price = self.ws[i][2].value

            if name in self.categories:
                self.categorie = name
                continue
            elif name is None or price == 0:
                continue
            if " Summer Cover" in name:
                continue
            if price is None:
                continue
            else:
                name = get_product_name(product_name=name, category=self.categorie)
                print(name)
                if os.path.exists(path):
                    product_parameters = RRP.create(name=name, price=int(price))
                else:
                    create_db_from_file.create_database()
                    product_parameters = RRP.create(name=name, price=int(price))

    def wb_active(self):
        wb = load_workbook(self.file_name, read_only=True)
        self.ws = wb.active


if __name__ == "__main__":
    fp = FileProcess("price.xlsx")
    fp.main()
