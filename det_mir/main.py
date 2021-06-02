from docx import Document
from openpyxl import load_workbook

from det_mir.create_tables import create_high_table,create_bottom_table

class FileProcessing:
    def __init__(self, file_name, export_file_name):
        self.file = file_name
        self.export_name = export_file_name
        self.supplier_name = None
        self.order_num = None
        self.total_quantity = None
        self.destination = None

    def main(self):

        wb = load_workbook(self.file)

        self.supplier_name = wb['Список'][2][0].value
        self.order_num = wb['Список'][2][1].value
        self.total_quantity = int(wb['Список'][2][5].value)
        self.destination = wb['Список'][2][8].value
        sheet = wb['Заказ']

        global_number = 1
        document = Document()
        for row in range(27, sheet.max_row + 1):
            product_quantity: int = int(sheet[row][10].value)
            product_name: str = sheet[row][3].value
            barcode: str = sheet[row][1].value
            model_number: str = sheet[row][2].value

            product_name = product_name.split('Joie')[-1]

            for product_position in range(1, product_quantity + 1):
                position = f'Место {global_number} из {self.total_quantity}'
                create_high_table(document, self.supplier_name, self.order_num, self.destination, position)
                if global_number % 3 != 0:
                    create_bottom_table(document, barcode, model_number, product_name, break_flag=True)
                else:
                    create_bottom_table(document, barcode, model_number, product_name, break_flag=False)
                global_number += 1
        document.save(f'{self.export_name}.docx')


if __name__ == '__main__':
    fp = FileProcessing('det_mir.xlsx')
    fp.main()
